"""Shared helpers for catalog & subcontractor item management:

  * generate_item_code — system-assigned, never-reused codes
  * summarize_changes / log_item_change — the edit/delete audit trail
  * build_items_workbook — export filled item data to an .xlsx sheet
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session

from .models import CatalogItem, Company, ItemAudit

# Fields shown in the audit "what changed" summary and the Excel export.
_TRACKED = [
    ("description_en", "Description (EN)"),
    ("description_ar", "Description (AR)"),
    ("unit", "Measure unit"),
    ("count_unit", "Count unit"),
    ("unit_cost", "Unit cost"),
    ("brand", "Brand"),
    ("industry", "Industry"),
    ("category", "Category"),
    ("supplier", "Supplier"),
    ("model_number", "Model / part no."),
    ("link", "Link"),
    ("notes", "Notes"),
]


# --- auto-generated, never-reused item codes --------------------------------


def _code_in_use(db: Session, company_id: int, code: str) -> bool:
    """True if the code is on a live item OR was ever used (in the audit log)."""
    live = db.execute(
        select(CatalogItem.id).where(
            CatalogItem.company_id == company_id, CatalogItem.item_code == code
        )
    ).first()
    if live:
        return True
    used = db.execute(
        select(ItemAudit.id).where(
            ItemAudit.company_id == company_id, ItemAudit.item_code == code
        )
    ).first()
    return bool(used)


def generate_item_code(db: Session, company: Company) -> str:
    """Assign the next per-company item code (e.g. ITM-000123). The counter only
    ever increases, so codes are never reused — not even after deletion.

    The increment is done with an atomic ``UPDATE ... RETURNING`` so two
    concurrent creates can never claim the same number (the row lock serialises
    them at the database). Skips any code that somehow already exists."""
    while True:
        new_seq = db.execute(
            update(Company)
            .where(Company.id == company.id)
            .values(next_item_seq=func.coalesce(Company.next_item_seq, 1) + 1)
            .returning(Company.next_item_seq)
        ).scalar_one()
        seq = new_seq - 1  # the value we just claimed (pre-increment)
        code = f"ITM-{seq:06d}"
        if not _code_in_use(db, company.id, code):
            return code


# --- audit trail ------------------------------------------------------------


def _fmt(value) -> str:
    if value is None or value == "":
        return "∅"
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def summarize_changes(item: CatalogItem, fields: dict) -> str:
    """Human-readable 'field: old → new' summary of an edit. Empty if nothing
    tracked actually changed."""
    parts = []
    for key, label in _TRACKED:
        if key not in fields:
            continue
        old = getattr(item, key)
        new = fields[key]
        if _fmt(old) != _fmt(new):
            parts.append(f"{label}: {_fmt(old)} → {_fmt(new)}")
    return "; ".join(parts)


def log_item_change(
    db: Session, item: CatalogItem, action: str, user, details: str = ""
) -> None:
    """Append an audit row for an edit/delete. Caller commits."""
    db.add(
        ItemAudit(
            company_id=item.company_id,
            subcontractor_id=item.subcontractor_id,
            item_code=item.item_code,
            item_description=item.description_en or item.description_ar,
            action=action,
            details=details or None,
            user_id=getattr(user, "id", None),
            username=getattr(user, "full_name", None) or getattr(user, "username", None),
        )
    )


# --- Excel export -----------------------------------------------------------

_EXPORT_COLUMNS = [
    ("item_code", "Item Code", 16),
    ("description_en", "Description (EN)", 34),
    ("description_ar", "Description (AR)", 34),
    ("unit", "Measure Unit", 14),
    ("count_unit", "Count Unit", 12),
    ("unit_cost", "Unit Cost", 12),
    ("brand", "Brand", 16),
    ("industry", "Industry", 16),
    ("category", "Category", 16),
    ("supplier", "Supplier", 18),
    ("model_number", "Model / Part No.", 18),
    ("link", "Link", 30),
    ("notes", "Notes", 30),
]
_HEADER_FILL = PatternFill("solid", fgColor="1A56DB")


def build_items_workbook(items: list[CatalogItem], title: str = "Catalog") -> bytes:
    """An .xlsx of the filled item data (one row per item)."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Catalog"

    for c, (_, header, width) in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 24

    for r, item in enumerate(items, start=2):
        for c, (attr, _, _) in enumerate(_EXPORT_COLUMNS, start=1):
            value = getattr(item, attr, None)
            if attr == "unit_cost":
                value = float(value or 0)
            cell = ws.cell(row=r, column=c, value=value)
            cell.alignment = Alignment(
                vertical="top", wrap_text=attr in ("description_en", "description_ar", "notes")
            )
            if attr == "unit_cost":
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Columns offered in the upload template. item_code is intentionally omitted —
# the system assigns codes automatically. Only unit_cost is required.
_TEMPLATE_COLUMNS = [
    ("description_en", "description_en", 32),
    ("description_ar", "description_ar", 32),
    ("unit", "unit", 12),
    ("count_unit", "count_unit", 12),
    ("unit_cost", "unit_cost", 12),
    ("brand", "brand", 16),
    ("industry", "industry", 16),
    ("category", "category", 16),
    ("supplier", "supplier", 18),
    ("model_number", "model_number", 16),
    ("link", "link", 28),
    ("notes", "notes", 26),
]
_TEMPLATE_EXAMPLES = [
    {
        "description_en": "LED panel light 60x60",
        "description_ar": "لوحة إضاءة LED ٦٠×٦٠",
        "unit": "Each",
        "count_unit": "each",
        "unit_cost": 42.5,
        "brand": "Philips",
        "industry": "Electrical",
        "category": "Lighting",
        "supplier": "Acme Supply",
        "model_number": "RC065B",
        "link": "https://example.com/datasheet.pdf",
        "notes": "4000K, 36W (example row — replace or delete)",
    },
    {
        "description_en": "Copper cable 3-core 2.5mm2",
        "description_ar": "كابل نحاس ٣ قلوب ٢.٥ مم٢",
        "unit": "m",
        "count_unit": "roll",
        "unit_cost": 4.7,
        "brand": "Nexans",
        "industry": "Electrical",
        "category": "Cables",
        "supplier": "Gulf Cables",
        "model_number": "",
        "link": "",
        "notes": "example row — replace or delete",
    },
]


def build_catalog_template() -> bytes:
    """A ready-to-fill .xlsx: headers the importer accepts + two example rows.
    Only `unit_cost` is mandatory; item codes are assigned automatically."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Catalog template"

    for c, (_, header, width) in enumerate(_TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 22

    for r, example in enumerate(_TEMPLATE_EXAMPLES, start=2):
        for c, (key, _, _) in enumerate(_TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=r, column=c, value=example.get(key))
            cell.alignment = Alignment(vertical="top", wrap_text=key.startswith("description") or key == "notes")
            if key == "unit_cost":
                cell.number_format = "#,##0.00"

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

"""Build a bilingual (AR/EN) Bill of Quantities as an .xlsx workbook.

Pure formatting — given the RFP document and its scope lines with the BoQ lines
to include, returns the workbook bytes. The router decides which lines to pass
(approved-only by default).
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# (header text — EN over AR, column width)
COLUMNS = [
    ("Item Code\nرمز البند", 16),
    ("Description (EN)\nالوصف (إنجليزي)", 36),
    ("Description (AR)\nالوصف (عربي)", 36),
    ("Unit\nالوحدة", 10),
    ("Qty\nالكمية", 10),
    ("Unit Price\nسعر الوحدة", 14),
    ("Total\nالإجمالي", 16),
    ("Brand / Comments\nالعلامة / ملاحظات", 22),
]
NCOLS = len(COLUMNS)
MONEY_FMT = "#,##0.00"

_HEADER_FILL = PatternFill("solid", fgColor="1A56DB")
_SECTION_FILL = PatternFill("solid", fgColor="14202E")
_GROUP_FILL = PatternFill("solid", fgColor="EEF2FF")
_TOTAL_FILL = PatternFill("solid", fgColor="F3F4F6")
_thin = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _f(value) -> float:
    return float(value or 0)


def build_boq_workbook(filename: str, groups: list) -> bytes:
    """groups: list of (scope_line, [boq_line, ...]) — only non-empty groups."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BoQ"

    # Title + meta
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    title = ws.cell(row=1, column=1, value="Bill of Quantities — جدول الكميات")
    title.font = Font(bold=True, size=16)
    title.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NCOLS)
    meta = ws.cell(
        row=2,
        column=1,
        value=f"Source: {filename}    Generated: {datetime.now():%Y-%m-%d %H:%M}",
    )
    meta.font = Font(size=9, color="666666")
    meta.alignment = Alignment(horizontal="center")

    # Header row
    header_row = 4
    for c, (text, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=c, value=text)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[header_row].height = 30

    row = header_row + 1
    grand_total = 0.0

    # Group scope lines by SECTION (from AI analysis): the section becomes the
    # main "Scope N" heading, and the items/tasks are listed under it. Plain
    # table uploads (no sections) render as a single flat list.
    from collections import OrderedDict

    sections: "OrderedDict[int, dict]" = OrderedDict()
    for scope, lines in groups:
        sno = getattr(scope, "section_no", 0) or 0
        if sno not in sections:
            sections[sno] = {"title": getattr(scope, "section_title", None), "groups": []}
        elif not sections[sno]["title"]:
            sections[sno]["title"] = getattr(scope, "section_title", None)
        sections[sno]["groups"].append((scope, lines))

    has_sections = not (len(sections) == 1 and 0 in sections and not sections[0]["title"])

    for display_i, sec in enumerate(sections.values(), start=1):
        if has_sections:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
            label = f"Scope {display_i}" + (f" — {sec['title']}" if sec["title"] else "")
            hcell = ws.cell(row=row, column=1, value=label)
            hcell.font = Font(bold=True, color="FFFFFF", size=11)
            hcell.fill = _SECTION_FILL
            hcell.alignment = Alignment(horizontal="left")
            for c in range(1, NCOLS + 1):
                ws.cell(row=row, column=c).border = _BORDER
            row += 1

        section_total = 0.0
        for scope, lines in sec["groups"]:
            for bl in lines:
                # Use the concise catalog description; for an unmatched line fall
                # back to the (summarized) task text from the scope line.
                desc_en = bl.description_en or (scope.description if not bl.item_code else "")
                brand_cell = " · ".join(
                    p for p in (bl.brand, getattr(bl, "subcontractor", None)) if p
                )
                values = [
                    bl.item_code or "—",
                    desc_en or "",
                    bl.description_ar or "",
                    bl.unit or "",
                    _f(bl.quantity),
                    _f(bl.unit_price),
                    _f(bl.line_total),
                    brand_cell,
                ]
                for c, v in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.border = _BORDER
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=c in (2, 3),
                        horizontal="right" if c in (3, 5, 6, 7) else "left",
                    )
                    if c in (6, 7):
                        cell.number_format = MONEY_FMT
                section_total += _f(bl.line_total)
                row += 1

        # Section subtotal
        scell = ws.cell(row=row, column=6, value="Subtotal — مجموع جزئي")
        scell.font = Font(bold=True, italic=True, size=9)
        scell.alignment = Alignment(horizontal="right")
        tcell = ws.cell(row=row, column=7, value=round(section_total, 2))
        tcell.font = Font(bold=True)
        tcell.number_format = MONEY_FMT
        for c in range(1, NCOLS + 1):
            ws.cell(row=row, column=c).border = _BORDER
        row += 1
        grand_total += section_total

    # Grand total
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    gt_label = ws.cell(row=row, column=1, value="GRAND TOTAL — الإجمالي الكلي")
    gt_label.font = Font(bold=True, size=12)
    gt_label.alignment = Alignment(horizontal="right")
    gt_label.fill = _TOTAL_FILL
    gt_val = ws.cell(row=row, column=7, value=round(grand_total, 2))
    gt_val.font = Font(bold=True, size=12)
    gt_val.number_format = MONEY_FMT
    gt_val.fill = _TOTAL_FILL
    for c in range(1, NCOLS + 1):
        ws.cell(row=row, column=c).border = _BORDER

    ws.freeze_panes = "A5"  # keep title + header visible while scrolling

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

"""Parse and validate a supplier rate sheet (CSV or XLSX) into catalog rows.

Pure parsing/validation — no database access here, so it's easy to test and
reason about. The router calls `parse_catalog_file()` and decides what to do
with the results.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

# Canonical column -> set of accepted header spellings (all matched after
# normalisation: lowercased, trimmed, spaces/hyphens collapsed to underscore).
HEADER_ALIASES: dict[str, set[str]] = {
    "item_code": {"item_code", "code", "sku", "item_no", "item_number"},
    "description_ar": {"description_ar", "arabic_description", "arabic", "ar", "desc_ar"},
    "description_en": {"description_en", "english_description", "english", "en", "desc_en", "description"},
    "unit": {"unit", "uom", "units"},
    "material_cost": {"material_cost", "material", "mat_cost"},
    "labour_cost": {"labour_cost", "labor_cost", "labour", "labor"},
    "markup": {"markup", "markup_pct", "markup_percent", "margin"},
    "brand": {"brand", "manufacturer", "make", "maker"},
    # Advanced optional columns.
    "industry": {"industry", "trade", "discipline", "sector"},
    "category": {"category", "subcategory", "sub_category", "type", "group"},
    "supplier": {"supplier", "vendor", "distributor", "source"},
    "model_number": {"model_number", "model", "part_number", "part_no", "mpn", "model_no"},
    "link": {"link", "url", "links", "reference", "reference_url", "datasheet", "spec_url"},
    "notes": {"notes", "note", "remarks", "comment", "comments", "specs", "specification"},
}

REQUIRED_COLUMNS = {"item_code", "unit", "material_cost", "labour_cost", "markup"}
NUMERIC_COLUMNS = {"material_cost", "labour_cost", "markup"}


@dataclass
class ParseResult:
    valid_rows: list[dict] = field(default_factory=list)
    row_errors: list[tuple[int, list[str]]] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)
    duplicates_in_file: int = 0


def _normalize_header(h: str) -> str:
    h = (h or "").strip().lower()
    for ch in (" ", "-", "/"):
        h = h.replace(ch, "_")
    while "__" in h:
        h = h.replace("__", "_")
    return h.strip("_")


def _build_header_map(raw_headers: list[str]) -> dict[int, str]:
    """Map each source column index -> canonical name (unknown columns dropped)."""
    lookup = {alias: canon for canon, aliases in HEADER_ALIASES.items() for alias in aliases}
    col_map: dict[int, str] = {}
    for idx, raw in enumerate(raw_headers):
        canon = lookup.get(_normalize_header(raw))
        if canon and canon not in col_map.values():
            col_map[idx] = canon
    return col_map


def _read_table(filename: str, content: bytes) -> list[list[str]]:
    """Return rows (incl. header) as lists of stringified cells."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        # utf-8-sig strips a BOM if Excel added one; keeps Arabic intact.
        text = content.decode("utf-8-sig", errors="replace")
        return [row for row in csv.reader(io.StringIO(text))]
    if name.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows: list[list[str]] = []
        for r in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c).strip() for c in r])
        wb.close()
        return rows
    raise ValueError("Unsupported file type. Upload a .csv or .xlsx file.")


def _parse_number(value: str) -> float:
    v = (value or "").strip().replace(",", "")
    if v == "":
        return 0.0
    return float(v)  # raises ValueError on garbage; caller catches it


def parse_catalog_file(filename: str, content: bytes) -> ParseResult:
    table = _read_table(filename, content)
    result = ParseResult()

    # Drop fully-empty leading rows, then take the first non-empty row as header.
    rows = [r for r in table if any((c or "").strip() for c in r)]
    if not rows:
        result.missing_columns = sorted(REQUIRED_COLUMNS)
        return result

    header, data_rows = rows[0], rows[1:]
    col_map = _build_header_map(header)
    present = set(col_map.values())
    result.missing_columns = sorted(REQUIRED_COLUMNS - present)
    if result.missing_columns:
        return result  # caller returns 400; nothing else to do

    seen_codes: dict[str, int] = {}  # item_code -> index in valid_rows
    for i, raw in enumerate(data_rows, start=1):
        record: dict = {c: None for c in HEADER_ALIASES}
        errors: list[str] = []

        for idx, canon in col_map.items():
            cell = raw[idx].strip() if idx < len(raw) else ""
            if canon in NUMERIC_COLUMNS:
                try:
                    record[canon] = _parse_number(cell)
                except ValueError:
                    errors.append(f"{canon}: '{cell}' is not a number")
            else:
                record[canon] = cell or None

        if not record["item_code"]:
            errors.append("item_code is required")
        if not record["description_ar"] and not record["description_en"]:
            errors.append("at least one of description_ar / description_en is required")
        for col in NUMERIC_COLUMNS:
            if isinstance(record[col], float) and record[col] < 0:
                errors.append(f"{col} must be >= 0")

        if errors:
            result.row_errors.append((i, errors))
            continue

        code = record["item_code"]
        if code in seen_codes:  # later duplicate wins, overwrite earlier
            result.valid_rows[seen_codes[code]] = record
            result.duplicates_in_file += 1
        else:
            seen_codes[code] = len(result.valid_rows)
            result.valid_rows.append(record)

    return result
